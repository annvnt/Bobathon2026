"""Parse ECHA chemical list XLSX exports in ECHA/."""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

from radar.config import ECHA_DIR, PARTNERS_FILE
from radar.compliance import taxonomy

LIST_URLS: dict[str, str] = {
    "candidate_list": "https://echa.europa.eu/candidate-list-table",
    "restriction_list": "https://echa.europa.eu/substances-restricted-under-reach",
    "authorisation_list": "https://echa.europa.eu/authorisation-list",
    "eu_positive_list": "https://echa.europa.eu/eu-positive-list",
    "svhc_table": "https://echa.europa.eu/candidate-list-table",
}

# Official ECHA full exports (preferred over third-party duplicates).
PREFERRED_FILES = {
    "candidate_list": "candidate_list_full-2026-02-27.xlsx",
    "restriction_list": "restriction_list_full-2026-04-29.xlsx",
    "authorisation_list": "authorisation_list_full-2025-09-13.xlsx",
    "eu_positive_list": "eu_positive_list_full-2025-12-11.xlsx",
}

DATE_IN_NAME = re.compile(r"(\d{4}-\d{2}-\d{2})")


def _parse_excel_date(val: Any) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    text = str(val).strip()
    for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _date_from_filename(path: Path) -> str:
    m = DATE_IN_NAME.search(path.stem)
    return m.group(1) if m else date.today().isoformat()


def _list_kind(path: Path) -> str | None:
    stem = path.stem.lower()
    if stem.startswith("candidate_list"):
        return "candidate_list"
    if stem.startswith("restriction_list"):
        return "restriction_list"
    if stem.startswith("authorisation_list"):
        return "authorisation_list"
    if stem.startswith("eu_positive_list"):
        return "eu_positive_list"
    if "svhc" in stem:
        return "svhc_table"
    return None


def _portfolio_substances() -> set[str]:
    import json

    if not PARTNERS_FILE.exists():
        return taxonomy.canonical_substances()
    data = json.loads(PARTNERS_FILE.read_text(encoding="utf-8"))
    found: set[str] = set()
    for partner in data.get("partners", []):
        for product in partner.get("products", []):
            found.update(product.get("substances") or [])
    return found or taxonomy.canonical_substances()


def _row_text(name: str, description: str, ec: str, cas: str) -> str:
    parts = [name, description, ec, cas]
    return " ".join(p for p in parts if p and p != "-")


def _normalize_cas(cas: str) -> str:
    return cas.strip().rstrip(",").strip()


def _parse_standard_row(
    row: tuple,
    headers: dict[str, int],
    list_kind: str,
    file_date: str,
    source_file: str,
) -> dict[str, Any] | None:
    def col(key: str) -> str:
        idx = headers.get(key)
        if idx is None or idx >= len(row):
            return ""
        val = row[idx]
        return "" if val is None else str(val).strip()

    name = col("substance name")
    if not name or name.lower() == "substance name":
        return None

    description = col("description")
    ec = col("ec number")
    cas = _normalize_cas(col("cas number"))

    text = _row_text(name, description, ec, cas)
    substances = taxonomy.resolve_substances(text)
    if not substances:
        return None

    portfolio = _portfolio_substances()
    if not (substances & portfolio):
        return None

    effective = None
    for date_col in ("date of inclusion", "regulatory outcome date", "sunset date", "expiry date"):
        if date_col in headers:
            effective = _parse_excel_date(row[headers[date_col]])
            if effective:
                break
    effective = effective or file_date

    entry_num = col("entry number") if "entry number" in headers else ""
    conditions = col("conditions") if "conditions" in headers else ""

    label = list_kind.replace("_", " ")
    ref_parts = [f"ECHA {label}", name[:120]]
    if cas and cas != "-":
        ref_parts.append(cas)
    elif ec and ec != "-":
        ref_parts.append(ec)

    return {
        "list_kind": list_kind,
        "name": name,
        "description": description if description != "-" else "",
        "ec_number": ec if ec != "-" else "",
        "cas_number": cas if cas != "-" else "",
        "substances": sorted(substances),
        "effective_date": effective,
        "file_date": file_date,
        "source_file": source_file,
        "reference": ": ".join(ref_parts[:2]) + (f" ({cas or ec})" if (cas or ec) not in ("", "-") else ""),
        "entry_number": entry_num,
        "conditions": conditions,
        "url": LIST_URLS.get(list_kind, LIST_URLS["candidate_list"]),
        "summary_extra": _summary_for_kind(list_kind),
    }


def _summary_for_kind(list_kind: str) -> str:
    return {
        "candidate_list": "SVHC Candidate List — Article 33 communication / SCIP obligations may apply.",
        "restriction_list": "REACH Annex XVII restriction — check concentration limits and conditions.",
        "authorisation_list": "REACH Annex XIV authorisation — sunset date may apply for uses.",
        "eu_positive_list": "EU positive list (food contact materials) — verify FCM compliance if applicable.",
        "svhc_table": "SVHC Candidate List (legacy export).",
    }.get(list_kind, "ECHA substance listing.")


def _header_map(header_row: tuple) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if key:
            mapping[key] = i
    return mapping


def _iter_standard_rows(path: Path, list_kind: str) -> Iterator[dict[str, Any]]:
    from openpyxl import load_workbook

    file_date = _date_from_filename(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers: dict[str, int] | None = None
    try:
        for row in ws.iter_rows(values_only=True):
            if not row or not any(row):
                continue
            if headers is None:
                first = str(row[0] or "").strip().lower()
                if first == "substance name" or first == "index":
                    headers = _header_map(row)
                    continue
                if first == "index" and str(row[1] or "").lower() == "chemical name":
                    headers = {
                        "index": 0,
                        "substance name": 1,
                        "ec number": 2,
                        "cas number": 3,
                        "date of inclusion": 4,
                    }
                    continue
                continue
            parsed = _parse_standard_row(row, headers, list_kind, file_date, path.name)
            if parsed:
                yield parsed
    finally:
        wb.close()


def _files_to_load(*, include_eu_positive: bool = False) -> list[Path]:
    if not ECHA_DIR.exists():
        return []
    paths: list[Path] = []
    for kind, preferred in PREFERRED_FILES.items():
        if kind == "eu_positive_list" and not include_eu_positive:
            continue
        p = ECHA_DIR / preferred
        if p.exists():
            paths.append(p)
    loaded = {p.name for p in paths}
    for p in sorted(ECHA_DIR.glob("*.xlsx")):
        kind = _list_kind(p)
        if not kind or p.name in loaded:
            continue
        if kind == "svhc_table" and any(_list_kind(x) == "candidate_list" for x in paths):
            continue
        if kind == "eu_positive_list":
            continue
        paths.append(p)
    return paths


def load_entries(*, include_eu_positive: bool = False, portfolio_only: bool = True) -> list[dict[str, Any]]:
    """
    Load substance entries from ECHA/ XLSX files.
    By default includes rows whose substances appear in the organizer portfolio.
    """
    entries: list[dict[str, Any]] = []
    seen: set[str] = set()

    for path in _files_to_load(include_eu_positive=include_eu_positive):
        kind = _list_kind(path)
        if not kind:
            continue
        if kind == "eu_positive_list" and not include_eu_positive:
            continue
        try:
            for row in _iter_standard_rows(path, kind):
                key = f"{kind}|{row.get('cas_number') or row.get('ec_number') or row['name']}"
                if key in seen:
                    continue
                seen.add(key)
                entries.append(row)
        except Exception as e:
            print(f"ECHA: skip {path.name} — {e}")

    if not portfolio_only:
        return entries
    return entries


def stats() -> dict[str, Any]:
    files = _files_to_load(include_eu_positive=False)
    return {
        "files": [p.name for p in files],
        "portfolio_substances": sorted(_portfolio_substances()),
    }
